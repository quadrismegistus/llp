from __future__ import absolute_import
from __future__ import print_function
import codecs,configparser,os,re
import six
from six.moves import range
from six.moves import zip
from functools import reduce
from smart_open import open

try:
    input = raw_input
except NameError:
    pass


from os.path import expanduser
HOME=expanduser("~")
LLP_ROOT = LIT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))

if not 'llp.py' in os.listdir(LLP_ROOT):
	LLP_ROOT = LIT_ROOT = os.path.join(LLP_ROOT,'llp')

if not 'llp.py' in os.listdir(LLP_ROOT):
	print('!?',LLP_ROOT)

#print('LLP root:',LIT_ROOT)



PATH_BASE_CONF=os.path.join(HOME,'.llp_config')
PATH_DEFAULT_CONF=os.path.abspath(os.path.join(LIT_ROOT,'..','config_default.txt'))

PATH_MANIFEST_GLOBAL = os.path.join(LIT_ROOT,'corpus','manifest.txt')
#print(PATH_MANIFEST_GLOBAL, os.path.exists(PATH_MANIFEST_GLOBAL))




### SET THE CONFIG

def hash(x):
	import hashlib
	return hashlib.sha224(str(x).encode('utf-8')).hexdigest()

### SET THE CONFIGS


def config_obj2dict(config_obj,keys=['Default','User'],pathhack_root=LIT_ROOT,pathhack=True):
	config_dict = {}

	#dict([(k.upper(),v) for k,v in list(config[key].items())])
	for key in keys:
		if key not in config_obj: continue
		for attr,val in config_obj[key].items():
			if pathhack and 'path' in attr.lower() and not os.path.isabs(val):
				val=os.path.abspath(os.path.join(os.path.dirname(pathhack_root), val))
			config_dict[attr.upper()]=val

	return config_dict


def load_config(pathhack=True,prompt_for_base_conf=True):
	if prompt_for_base_conf and not os.path.exists(PATH_BASE_CONF):
		configure_prompt()

	CONFIG={}
	for f in [load_default_config,load_global_config,load_user_config]:
		for k,v in f().items(): CONFIG[k.upper()]=v

	#print('>> loaded config:',CONFIG)
	return CONFIG



def load_global_config(pathhack=True,prompt_for_base_conf=True):
	#CONFIG_PATHS = [PATH_DEFAULT_CONF]
	CONFIG_PATHS=[]
	CONFIG_PATHS += [os.path.join(LIT_ROOT,'config_local.txt')]
	CONFIG_PATHS.append(os.path.join(os.path.join(LIT_ROOT,'..','llp_config.txt')))
	CONFIG_PATHS.append(os.path.join(os.path.join(LIT_ROOT,'..','config','llp_config.txt')))
	CONFIG_PATHS.append(os.path.join(os.path.join(HOME,'llp_config.txt')))

	CONFIG={}
	for config_path in CONFIG_PATHS:
		#print('## looking for config:',os.path.abspath(config_path))
		if not os.path.exists(config_path): continue
		config = configparser.ConfigParser()
		config.read(config_path)

		for k,v in config_obj2dict(config,pathhack_root=config_path).items():
			CONFIG[k]=v


	#print(CONFIG)
	return CONFIG


def load_default_config():
	config=configparser.ConfigParser()
	config.read(PATH_DEFAULT_CONF)
	return config_obj2dict(config,pathhack_root=LIT_ROOT)

def load_user_config():
	config=configparser.ConfigParser()
	if os.path.exists(PATH_BASE_CONF):
		with open(PATH_BASE_CONF) as f:
			path_base_conf_value = f.read().strip()
			if os.path.exists(path_base_conf_value):
				config.read(path_base_conf_value)
				configd=config_obj2dict(config,pathhack_root=path_base_conf_value)
				return configd
	return {}




def configure_prompt(default_config='~/llp/config.txt',default_corpora='~/llp/corpora',default_manifest='~/llp/manifest.txt'):
	print('## Literary Language Processing (LLP) configuration')

	path_config=input('>> Where should the config file be stored? [default: {default}]: '.format(default=default_config)).strip()
	path_corpora=input('>> Where should corpora be stored? [default: {default}]: '.format(default=default_corpora)).strip()
	path_manifest=input('>> Where should the corpus manifest be stored? [default: {default}] '.format(default=default_manifest)).strip()

	if not path_config: path_config=default_config
	if not path_corpora: path_corpora=default_corpora
	if not path_manifest: path_manifest=default_manifest

	path_config=path_config.replace('~',HOME)
	path_corpora=path_corpora.replace('~',HOME)
	path_manifest=path_manifest.replace('~',HOME)

	var2path = {}
	var2path['PATH_TO_CORPORA'] = path_corpora
	var2path['PATH_TO_MANIFEST'] = path_manifest

	for var,path in var2path.items():
		var2path[var] = path = path.replace('~',HOME)  #os.path.expanduser(path)

		# make dir if needed
		if not os.path.exists(path):
			if os.path.splitext(path)[0]==path:
				os.makedirs(path)
			else:
				dirname=os.path.dirname(path)
				if not os.path.exists(dirname):
					os.makedirs(dirname)

	config_obj = configparser.ConfigParser()

	newconfig={} #dict(load_config())
	#for k,v in load_default_config().items(): newconfig[k]=v
	for k,v in var2path.items(): newconfig[k]=v
	config_obj['User'] = newconfig

	with open(path_config,'w') as of:
		config_obj.write(of)
		print('>> saved:',path_config)

	with open(PATH_BASE_CONF,'w') as of:
		of.write(path_config)

	if not os.path.exists(path_manifest):
		import shutil
		shutil.copyfile(PATH_MANIFEST_GLOBAL,path_manifest)
		print('>> saved:',path_manifest)














# load config
config=load_config()




WORDDB_FN = config.get('PATH_TO_WORDDB')

ENGLISH=None

import sys
import csv
#csv.field_size_limit(sys.maxsize)

def get_stopwords(include_rank=None):
	STOPWORDS_PATH = config.get('PATH_TO_ENGLISH_STOPWORDS')
	if not STOPWORDS_PATH: raise Exception('!! PATH_TO_ENGLISH_STOPWORDS not set in config.txt')
	if not STOPWORDS_PATH.startswith(os.path.sep): STOPWORDS_PATH=os.path.join(LIT_ROOT,STOPWORDS_PATH)
	#sw1=set(codecs.open(STOPWORDS_PATH,encoding='utf-8').read().strip().split('\n'))
	sw1=set(open(STOPWORDS_PATH).read().strip().split('\n'))
	if include_rank and type(include_rank)==int:
		sw2={d['word'] for d in worddb() if int(d['rank'])<=include_rank}
		sw1|=sw2
	sw1|={'i','me','my','you','we','us','our','they','them','she','he','her','hers','his','him'}
	return sw1

def get_english_wordlist():
	ENG_PATH = config.get('PATH_TO_ENGLISH_WORDLIST')
	if not ENG_PATH: raise Exception('!! PATH_TO_ENGLISH_WORDLIST not set in config.txt')
	if not ENG_PATH.startswith(os.path.sep): ENG_PATH=os.path.join(LIT_ROOT,ENG_PATH)
	#print('loading english from %s' % ENG_PATH)
	#return set(codecs.open(ENG_PATH,encoding='utf-8').read().strip().split('\n'))
	return set(open(ENG_PATH).read().strip().split('\n'))

def get_spelling_modernizer():
	SPELLING_MODERNIZER_PATH = config.get('PATH_TO_ENGLISH_SPELLING_MODERNIZER')
	if not SPELLING_MODERNIZER_PATH: raise Exception('!! PATH_TO_ENGLISH_SPELLING_MODERNIZER not set in config.txt')
	if not SPELLING_MODERNIZER_PATH.startswith(os.path.sep): SPELLING_MODERNIZER_PATH=os.path.join(LIT_ROOT,SPELLING_MODERNIZER_PATH)

	#print('>> getting spelling modernizer from %s...' % SPELLING_MODERNIZER_PATH)
	d={}
	#with codecs.open(SPELLING_MODERNIZER_PATH,encoding='utf-8') as f:
	with open(SPELLING_MODERNIZER_PATH) as f:
		for ln in f:
			ln=ln.strip()
			if not ln: continue
			try:
				old,new=ln.split('\t')
			except ValueError:
				continue
			d[old]=new
	return d

def get_ocr_corrections():
	PATH_TO_ENGLISH_OCR_CORRECTION_RULES = config.get('PATH_TO_ENGLISH_OCR_CORRECTION_RULES')
	if not PATH_TO_ENGLISH_OCR_CORRECTION_RULES: raise Exception('!! PATH_TO_ENGLISH_OCR_CORRECTION_RULES not set in config.txt')
	if not PATH_TO_ENGLISH_OCR_CORRECTION_RULES.startswith(os.path.sep): PATH_TO_ENGLISH_OCR_CORRECTION_RULES=os.path.join(LIT_ROOT,PATH_TO_ENGLISH_OCR_CORRECTION_RULES)

	#print('>> getting corrections from %s...' % PATH_TO_ENGLISH_OCR_CORRECTION_RULES)
	d={}


	with open(PATH_TO_ENGLISH_OCR_CORRECTION_RULES) as f:
		for ln in f:
			ln=ln.strip()
			if not ln: continue
			try:
				old,new,count=ln.split('\t')
			except ValueError:
				continue
			d[old]=new
	return d

def iter_filename(fn,force=False):
	if os.path.exists(fn) or force:
		filename,ext = os.path.splitext(fn)
		fnum=2 if not force else 1
		while os.path.exists(filename + str(fnum) + ext):
			fnum+=1
		fn = filename + str(fnum) + ext
	return fn


def measure_ocr_accuracy(txt_or_tokens):
	global ENGLISH
	if type(txt_or_tokens) in [str,six.text_type]:
		tokens=tokenize(txt_or_tokens)
	elif type(txt_or_tokens) in [tuple,list]:
		tokens=list(txt_or_tokens)
	else:
		raise Exception("Function `measure_ocr_accuracy(txt_or_tokens)` must take text string or list of tokens.")

	if not ENGLISH: ENGLISH = get_english_wordlist()

	numwords=float(len(tokens))
	numenglish=len([tok for tok in tokens if tok in ENGLISH])
	return numenglish/numwords


def tokenize(txt):
	from nltk import word_tokenize
	return word_tokenize(txt)



def find_nth_character(str1, substr, n):
	pos = -1
	for x in range(n):
		pos = str1.find(substr, pos+1)
		if pos == -1:
			return None
	return pos

## only singular nouns!
def to_singular(ld):
	import inflect
	p=inflect.engine()
	return [d for d in ld if p.singular_noun(d['word']) in {d['word'],False}]

def worddf():
	WORDDB_PATH = config.get('PATH_TO_WORDDB')
	if not WORDDB_PATH: raise Exception('!! PATH_TO_WORDDB not set in config.txt')
	if not WORDDB_PATH.startswith(os.path.sep): WORDDB_PATH=os.path.join(LIT_ROOT,WORDDB_PATH)

	import pandas as pd
	return pd.read_csv(WORDDB_PATH,sep='\t')

def worddb(abs_key = 'Complex Substance (Locke) <> Mixed Modes (Locke)_max',conc_key='Complex Substance (Locke) <> Mixed Modes (Locke)_min',cutoff_abs=0.1,cutoff_conc=-0.1,allow_names=False,only_content_words=True):
	WORDDB_PATH = config.get('PATH_TO_WORDDB')
	if not WORDDB_PATH: raise Exception('!! PATH_TO_WORDDB not set in config.txt')
	if not WORDDB_PATH.startswith(os.path.sep): WORDDB_PATH=os.path.join(LIT_ROOT,WORDDB_PATH)


	worddb = read_ld(WORDDB_PATH)
	for d in worddb:
		d['Abstract/Concrete'] = ''

		abs_score = float(d[abs_key])
		conc_score = float(d[conc_key])
		if only_content_words and d['is_content_word']!='True': continue
		if not allow_names and d['is_name']=='True': continue

		if abs_score >= cutoff_abs:
			d['Abstract/Concrete'] = 'Abstract'
		elif conc_score <= cutoff_conc:
			d['Abstract/Concrete'] = 'Concrete'
		else:
			d['Abstract/Concrete'] = 'Neither'

	return worddb


###





def read_ld(fn,keymap={},toprint=True):
	if fn.endswith('.xls') or fn.endswith('.xlsx'):
		return xls2ld(fn,keymap=keymap)
	#elif fn.endswith('.csv'):
		#return DictReader
		#return tsv2ld(fn,tsep=',',keymap=keymap)
		#tsep=','
		#with codecs.open(fn,encoding='utf-8') as f:
		#	return list(csv.DictReader(f))
	#else:
	#	tsep='\t'

	return list(readgen(fn,as_dict=True,toprint=toprint))


def writegen_jsonl(fnfn,generator,args=[],kwargs={}):
	import jsonlines
	with jsonlines.open(fnfn,'w') as writer:
		for i,dx in enumerate(generator(*args,**kwargs)):
			writer.write(dx)
	print('>> saved:',fnfn)

def readgen_jsonl(fnfn):
	import jsonlines
	with jsonlines.open(fnfn) as reader:
		for dx in reader:
			yield dx


def writegen(fnfn,generator,header=None,args=[],kwargs={}):
	import codecs,csv
	if 'jsonl' in fnfn.split('.'): return writegen_jsonl(fnfn,generator,args=args,kwargs=kwargs)
	iterator=generator(*args,**kwargs)
	first=next(iterator)
	if not header: header=sorted(first.keys())
	with open(fnfn, 'w') as csvfile:
		writer = csv.DictWriter(csvfile,fieldnames=header,extrasaction='ignore',delimiter='\t')
		writer.writeheader()
		for i,dx in enumerate(iterator):
			for k,v in dx.items():
				#if type(v) in [str]:
				#	dx[k]=v.encode('utf-8')
				dx[k] = str(v).replace('\r\n',' ').replace('\r',' ').replace('\n',' ').replace('\t',' ')
			writer.writerow(dx)
	print('>> saved:',fnfn)

def writegen_orig(fnfn,generator,header=None,args=[],kwargs={}):
	if 'jsonl' in fnfn.split('.'): return writegen_jsonl(fnfn,generator,args=args,kwargs=kwargs)
	with codecs.open(fnfn,'w',encoding='utf-8') as of:
		for i,dx in enumerate(generator()):
			if not header: header=sorted(dx.keys())
			if not i: of.write('\t'.join(header) + '\n')
			of.write('\t'.join([six.text_type(dx.get(h,'')) for h in header]) + '\n')
	print('>> saved:',fnfn)

def writegengen(fnfn,generator,header=None,save=True):
	if save: of = codecs.open(fnfn,'w',encoding='utf-8')
	for dx in generator():
		if not header:
			header=sorted(dx.keys())
			if save: of.write('\t'.join(header) + '\n')
		if save: of.write('\t'.join([six.text_type(dx.get(h,'')) for h in header]) + '\n')
		yield dx

def readgen_csv(fnfn,sep='\t',encoding='utf-8',errors='ignore'):
	#
	if fnfn.endswith('.gz'):
		from xopen import xopen
		f=xopen(fnfn)
	else:
		f=open(fnfn,encoding=encoding,errors=errors)

	reader = csv.DictReader(f,delimiter=sep,quoting=csv.QUOTE_NONE)
	for dx in reader:
		#for k,v in dx.items():
		#	if type(v)==str:
		#		dx[k]=v.decode(encoding=encoding, errors=errors)
		yield dx

	f.close()

def readgen(fnfn,header=None,tsep='\t',keymap={},keymap_all=six.text_type,encoding='utf-8',as_list=False,as_tuples=False,as_dict=True,toprint=True,progress=True):
	if 'jsonl' in fnfn.split('.'):
		for dx in readgen_jsonl(fnfn):
			yield dx
	else:
		import time
		now=time.time()

		if tsep=='\t' and toprint:
			print('>> streaming as tsv:',fnfn)
		elif tsep==',' and toprint:
			print('>> streaming as csv:',fnfn)

		if progress:
			num_lines = get_num_lines(fnfn)
			from tqdm import tqdm
			for dx in tqdm(readgen_csv(fnfn),total = num_lines): yield dx
		else:
			for dx in readgen_csv(fnfn): yield dx

		nownow=time.time()
		if not progress and toprint: print('   done ['+str(round(nownow-now,1))+' seconds]')

def header(fnfn,tsep='\t',encoding='utf-8'):
	header=[]

	if fnfn.endswith('.gz'):
		import gzip
		of=gzip.open(fnfn)
	#of = codecs.open(fnfn,encoding=encoding)
	else:
		of=open(fnfn)

	for line in of:
		line = line[:-1]  # remove line end character
		line=line.decode(encoding=encoding)
		header=line.split(tsep)
		break
	of.close()
	return header

# def read(fnfn,to_unicode=True):
# 	if fnfn.endswith('.gz'):
# 		import gzip
# 		try:
# 			with gzip.open(fnfn,'rb') as f:
# 				x=f.read()
# 				if to_unicode: x=x.decode('utf-8')
# 				return x
# 		except IOError as e:
# 			print("!! error:",e, end=' ')
# 			print("!! opening:",fnfn)
# 			print()
# 			return ''
#
# 	elif fnfn.endswith('.txt'):
# 		if to_unicode:
# 			try:
# 				with codecs.open(fnfn,encoding='utf-8') as f:
# 					return f.read()
# 			except UnicodeDecodeError:
# 				return read(fnfn,to_unicode=False)
# 		else:
# 			with open(fnfn) as f:
# 				return f.read()
#
# 	return ''

def read(fnfn):
	try:
		if fnfn.endswith('.gz'):
			import gzip
			with gzip.open(fnfn,'rb') as f:
				return f.read().decode('utf-8')
		else:
			with open(fnfn) as f:
				return f.read()
	except IOError as e:
		print("!! error:",e, end=' ')
		print("!! opening:",fnfn)
		print()
		return ''

def filesize(fn):
	return sizeof_fmt(os.path.getsize(fn))

def sizeof_fmt(num, suffix='B'):
	for unit in ['','Ki','Mi','Gi','Ti','Pi','Ei','Zi']:
		if abs(num) < 1024.0:
			return "%3.1f%s%s" % (num, unit, suffix)
		num /= 1024.0
	return "%.1f%s%s" % (num, 'Yi', suffix)



def xls2ld(fn,header=[],sheetname=True,keymap={},keymap_all=six.text_type):
	import time
	now=time.time()
	print('>> reading as xls:',fn)
	import xlrd
	if '*' in keymap: keymap_all=keymap['*']
	headerset=True if len(header) else False
	f=xlrd.open_workbook(fn)
	ld=[]
	def _boot_xls_sheet(sheet,header=[]):
		ld2=[]
		for y in range(sheet.nrows):
			if not header:
				for xi in range(sheet.ncols):
					cell=sheet.cell_value(rowx=y,colx=xi)
					header+=[cell]
				continue
			d={}
			for key in header:
				try:
					value=sheet.cell_value(rowx=y, colx=header.index(key))
					#print '??',value,type(value),key
					if keymap_all:
						func=keymap_all
						if func in [str,six.text_type] and type(value) in [float]:
							if value == int(value): value=int(value)
						d[key]=keymap_all(value)
					elif keymap and key in keymap:
						func=keymap[key]
						if func in [str,six.text_type] and type(value) in [float]:
							if value == int(value): value=int(value)
						d[key]=keymap[key](value)
					else:
						d[key]=value
					#print key,value,y,header.index(key),row[header.index(key)]
				except Exception as e:
					print('!! ERROR:',e)
					print('!! on key =',key,'& value =',value, type(value))
					#print "!! "+key+" not found in "+str(sheet)
					#d[key]=''
					pass
			if len(d):
				if sheetname: d['sheetname']=sheet.name
				ld2.append(d)
		return ld2


	if f.nsheets > 1:
		sheetnames=sorted(f.sheet_names())
		for sheetname in sheetnames:
			sheet=f.sheet_by_name(sheetname)
			for d in _boot_xls_sheet(sheet,header=header if headerset else []):
				ld.append(d)
	else:
		sheet = f.sheet_by_index(0)
		ld.extend(_boot_xls_sheet(sheet,header=header if headerset else []))

	nownow=time.time()
	print('>> done ['+str(round(nownow-now,1))+' seconds]')

	return ld


def xls2dld(fn,header=[]):
	return ld2dld(xls2ld(fn,header=header,sheetname=True), 'sheetname')

def levenshtein(s1, s2):
	l1 = len(s1)
	l2 = len(s2)

	matrix = [list(range(l1 + 1))] * (l2 + 1)
	for zz in range(l2 + 1):
		matrix[zz] = list(range(zz,zz + l1 + 1))
	for zz in range(0,l2):
		for sz in range(0,l1):
			if s1[sz] == s2[zz]:
				matrix[zz+1][sz+1] = min(matrix[zz+1][sz] + 1, matrix[zz][sz+1] + 1, matrix[zz][sz])
			else:
				matrix[zz+1][sz+1] = min(matrix[zz+1][sz] + 1, matrix[zz][sz+1] + 1, matrix[zz][sz] + 1)
	return matrix[l2][l1]

def xlsx2ld(fn,header=[],numsheets=1):
	from openpyxl.reader.excel import load_workbook
	header_set=bool(len(header))
	wb=load_workbook(filename=fn)
	ld=[]
	for sheet in wb.worksheets[:numsheets]:
		if not header_set: header=[]
		#header=[]
		for rownum,row in enumerate(sheet.rows):
			values=[]
			for cell in row:
				value=cell.value
				if value is None:
					value=''

				try:
					value=float(value)/0
				except:
					value=value
					if not isinstance(value, six.text_type):
						value=six.text_type(value)
				values.append(value)
			if not rownum and not len(header):
				header=values
			else:
				d=dict((header[i],values[i]) for i in range(len(values)))
				ld+=[d]
	return ld

def dl2ld(dl,kcol='group'):
	ld=[]
	for k in dl:
		for d in dl[k]:
			d[kcol]=k
			ld+=[d]
	return ld

def ld2dl(ld):
	keys = list(ld[0].keys())
	dl={}
	for k in keys:
		dl[k] = [d[k] for d in ld]
	return dl

def fn2ld(fn,header=[],sep='\t',nsep='\n'):
	import codecs
	f=codecs.open(fn,encoding='utf-8')
	for line in f:
		line=line.strip()
		if not header:
			header=line.split(sep)
			continue
		dx={}
		for i,val in enumerate(line.split(sep)):
			key=header[i] if len(header)>i else 'key_'+str(i)
			dx[key]=val
		yield dx

def goog2tsv(googsrc):
	import bs4
	dom=bs4.BeautifulSoup(googsrc,'html.parser')
	header=[th.text for th in dom('thead')[0]('th')]
	header=header if True in [bool(hx) for hx in header] else None
	old=[]
	for row in dom('tbody')[0]('tr'):
		rowdat=[cell.text for cell in row('td')]
		if not header:
			header=rowdat
			#print ">> HEADER:",header
			continue
		odx=dict(list(zip(header,rowdat)))
		old+=[odx]
	return old


def tsv2ld(fn,tsep='\t',nsep='\n',u=True,header=[],keymap={},zero='',removeEmpties=False):
	import time
	now=time.time()
	if tsep=='\t':
		print('>> reading as tsv:',fn)
	elif tsep==',':
		print('>> reading as csv:',fn)

	import os
	if fn.startswith('http'):
		print('>> reading webpage...')
		import urllib
		f=urllib.urlopen(fn)
		t=f.read()
		if fn.endswith('/pubhtml'):
			return goog2tsv(t)
		f.close()
	elif not os.path.exists(fn):
		t=fn
	elif u:
		import codecs
		f=codecs.open(fn,encoding='utf-8')
		t=f.read()
		f.close()
	else:
		f=open(fn,'r')
		t=f.read()
		f.close()
	t=t.replace('\r\n','\n')
	t=t.replace('\r','\n')

	#header=[]
	listdict=[]


	for line in t.split(nsep):
		if not line.strip(): continue
		line=line.replace('\n','')
		ln=line.split(tsep)
		#print ln
		if not header:
			header=ln
			for i,v in enumerate(header):
				if v.startswith('"') and v.endswith('"'):
					header[i]=v[1:-1]
			continue
		edict={}
		for i in range(len(ln)):
			try:
				k=header[i]
			except IndexError:
				#print "!! unknown column for i={0} and val={1}".format(i,ln[i])
				continue
			v=ln[i].strip()

			if '*' in keymap:
				v=keymap['*'](v)
			elif k in keymap:
				#print v, type(v)
				v=keymap[k](v)
				#print v, type(v)
			else:
				if v.startswith('"') and v.endswith('"'):
					v=v[1:-1]
				try:
					v=float(v)
				except ValueError:
					v=v

			if type(v) in [str,six.text_type] and not v:
				if zero=='' and removeEmpties:
					continue
				else:
					v=zero
			edict[k]=v
		if edict:
			listdict.append(edict)

	nownow=time.time()
	print('>> done ['+str(round(nownow-now,1))+' seconds]')

	return listdict



def ld2html(ld):
	keys=ld2keys(ld)
	headerrow=['<th>%s</th>'%k for k in keys]
	rows=[]
	rows+=['\n\t\t'.join(headerrow)]
	for d in ld:
		row=['<td>%s</td>'%d.get(k,'') for k in keys]
		rows+=['\n\t\t'.join(row)]
	ostr=u"<table>\n\t<tr>\n\t\t" + u'\n\t</tr>\n\t<tr>\n\t\t'.join(rows) + u"\n\t</tr>\n</table>"
	return ostr

def ld2keys(ld):
	keys=[]
	for d in ld:
		for k in d:
			keys+=[k]
	keys=list(sorted(list(set(keys))))
	return keys

def ld2ll(ld,zero='',tostr=False,uni=True):
	keys=[]
	for d in ld:
		for k in d:
			keys+=[k]
	keys=sorted(list(set(keys)))
	o=[keys]
	for d in ld:
		l=[]
		for k in keys:
			v=d.get(k,zero)
			if tostr:
				v=six.text_type(v) if uni else str(v)
			l+=[v]
		o+=[l]
	return o


def write_ld(fn,ld,zero='',timestamp=None):
	return write(fn,ld2ll(ld,zero=zero),timestamp=timestamp)

def dd2ld(dd,rownamecol='rownamecol'):
	if not rownamecol:
		return [ (dict(list(v.items()))) for k,v in list(dd.items()) ]
	else:
		return [ (dict(list(v.items()) + [(rownamecol,k)])) for k,v in list(dd.items()) ]

def dld2ld(dld,key='rownamecol'):
	ld=[]
	for k in dld:
		for d in dld[k]:
			d[key]=k
			ld+=[d]
	return ld

def ld_resample(ld,key='rownamecol',n=None):
	import random
	dld=ld2dld(ld,key)
	minlen_actually=min([len(dld[k]) for k in dld])
	minlen=minlen_actually if not n or n>minlen_actually else n
	ld2=[]
	print('>> resampling to minimum length of:',minlen)
	for k in sorted(dld):
		print('>>',k,len(dld[k]),'-->',minlen)
		ld2+=random.sample(dld[k],minlen)
	return ld2

def ld2dld(ld,key='rownamecol'):
	dld={}
	for d in ld:
		if not d[key] in dld: dld[d[key]]=[]
		dld[d[key]]+=[d]
	return dld

def ld2dd(ld,rownamecol='rownamecol'):
	dd={}
	for d in ld:
		dd[d[rownamecol]]=d
		#del dd[d[rownamecol]][rownamecol]
	return dd

def datatype(data,depth=0,v=False):
	def echo(dt):
		if not v: return
		for n in range(depth): print("\t", end=' ')
		print('['+dt[0]+']'+dt[1:], end=' ')
		try:
			print("[{0} records]".format(len(data),dt))
		except:
			print()

	if type(data) in [str,six.text_type]:
		echo('string')
		return 's'
	elif type(data) in [float,int]:
		echo('number')
		return 'n'
	elif type(data) in [list]:
		echo('list')
		if not len(data):
			return 'l'
		else:
			return 'l'+datatype(data[0],depth=depth+1,v=v)
	elif type(data) in [dict]:
		echo('dictionary')
		if not len(data):
			return 'd'
		else:
			return 'd'+datatype(list(data.values())[0],depth=depth+1,v=v)
	else:
		#print "WHAT TYPE OF DATA IS THIS:"
		#print data
		#print type(data)
		#print
		return '?'


def limcols(ld,limcol=255):
	keyd={}
	keys=set()
	for d in ld:
		dkeys=set(d.keys())
		for key in dkeys-keys:
			keyd[key]=0
		keys|=dkeys
		for k in d:
			if d[k]:
				keyd[k]+=1

	cols=set(sorted(list(keyd.keys()), key=lambda _k: (-keyd[_k],_k))[:limcol])

	for d in ld:
		dkeys=set(d.keys())
		for key in dkeys-cols:
			del d[key]

	return ld

def ld2str(ld,**data):
	if data['limcol']:
		print(">> limiting columns")
		limcol=data['limcol']
		ld=limcols(ld,limcol)
	if 'limcol' in data:
		del data['limcol']
	return ll2str(ld2ll(ld),**data)

def d2ll(d):
	try:
		return [[k,v] for k,v in sorted(list(d.items()),key=lambda lt: -lt[1])]
	except:
		return [[k,v] for k,v in list(d.items())]

def d2str(d,uni=True):
	return ll2str(d2ll(d),uni=uni)

def strmake(x,uni=True):
	if uni and type(x) in [six.text_type]:
		return x
	elif uni and type(x) in [str]:
		return x.decode('utf-8',errors='replace')
	elif uni:
		return six.text_type(x)
	elif not uni and type(x) in [str]:
		return x
	elif not uni and type(x) in [six.text_type]:
		return x.encode('utf-8',errors='replace')

	print([x],type(x))
	return str(x)


def ll2str(ll,uni=True,join_line=u'\n',join_cell=u'\t'):
	if not uni:
		join_line=str(join_line)
		join_cell=str(join_cell)
		quotechar='"' if join_cell==',' else ''
	else:
		quotechar=u'"' if join_cell==',' else u''

	for line in ll:
		lreturn=join_cell.join([quotechar+strmake(cell,uni=uni)+quotechar for cell in line])+join_line
		yield lreturn

def l2str(l,uni=True,join_line=u'\n',join_cell=u'\t',quotechar=''):
	for line in l: yield strmake(line)+join_line

def write_ld2(fn,gen1,gen2,uni=True,badkeys=[]):
	def find_keys(gen):
		keys=set()
		for d in gen:
			keys=keys|set(d.keys())
		keys=keys-set(badkeys)
		return keys

	keys=list(sorted(list(find_keys(gen1))))
	numk=len(keys)

	import codecs
	of=codecs.open(fn,'w',encoding='utf-8')
	of.write('\t'.join([strmake(x) for x in keys]) + '\n')

	for d in gen2:
		data=[d.get(key,'') for key in keys]
		of.write('\t'.join([strmake(x) for x in data]) + '\n')
	of.close()
	print(">> saved:",fn)


def write2(fn,data,uni=True,join_cell=u'\t',join_line=u'\n',limcol=None,toprint=True):
	## pass off to other write functions if necessary
	if fn.endswith('.xls'): return write_xls(fn,data)
	if fn.endswith('.csv'): join_cell=','

	## get datatyoe
	dt=datatype(data)

	## get str output for datatype
	if dt.startswith('ld'):
		o=ld2str(data,join_cell=join_cell,limcol=limcol)
	elif dt.startswith('dl'):
		o=dl2str(data,uni=uni)
	elif dt.startswith('ll'):
		o=ll2str(data,uni=uni)
	elif dt.startswith('dd'):
		o=dd2str(data,uni=uni)
	elif dt.startswith('l'):
		o=l2str(data,uni=uni)
	elif dt.startswith('d'):
		o=d2str(data,uni=uni)
	else:
		o=data

	## write
	import codecs
	of = codecs.open(fn,'w',encoding='utf-8') if True else open(fn,'w')
	for line in o: of.write(line)
	of.close()
	if toprint: print('>> saved:',fn)

def slice(l,num_slices=None,slice_length=None,runts=True,random=False):
	"""
	Returns a new list of n evenly-sized segments of the original list
	"""
	if random:
		import random
		random.shuffle(l)
	if not num_slices and not slice_length: return l
	if not slice_length: slice_length=int(len(l)/num_slices)
	newlist=[l[i:i+slice_length] for i in range(0, len(l), slice_length)]
	if runts: return newlist
	return [lx for lx in newlist if len(lx)==slice_length]


def noPunc(token):
	from string import punctuation
	return token.strip(punctuation)

def zeroPunc(s):
	import string
	return s.translate(str.maketrans('', '', string.punctuation))

def now(now=None):
	import datetime as dt
	if not now:
		now=dt.datetime.now()
	elif type(now) in [int,float,str]:
		now=dt.datetime.fromtimestamp(now)

	return '{0}-{1}-{2} {3}:{4}:{5}'.format(now.year,str(now.month).zfill(2),str(now.day).zfill(2),str(now.hour).zfill(2),str(now.minute).zfill(2),str(now.second).zfill(2))

def slingshot_cmd_starter(corpus,method,slingshot_n,slingshot_opts):
	Scmd='slingshot -llp_corpus {corpus} -llp_method {method}'.format(corpus=corpus,method=method)
	if slingshot_n: Scmd+=' -parallel {slingshot_n}'.format(slingshot_n=slingshot_n)
	if slingshot_opts: Scmd+=' '+slingshot_opts.strip()
	return Scmd



def toks2str(tlist,uni=False):
	toks=[]
	putleft=False
	#print tlist
	for tk in tlist:
		tk=tk.strip()
		if not tk: continue
		tk = tk.split()[-1]
		if not tk: continue
		if (not len(toks)):
			toks+=[tk]
		elif putleft:
			toks[-1]+=tk
			putleft=False
		elif tk=='`':
			toks+=[tk]
			putleft=True
		elif tk=='-LRB-':
			toks+=['(']
			putleft=True
		elif tk=='-RRB-':
			toks[-1]+=')'
		elif len(tk)>1 and tk[0]=="'":
			toks[-1]+=tk
		elif tk[0].isalnum():
			toks+=[tk]
		elif tk.startswith('<') and '>' in tk:
			toks+=[tk]
		else:
			toks[-1]+=tk
	if uni: return u' '.join(toks)
	return ' '.join(toks)






####
def print_config(corpus):
	print()
	print()
	print('[%s]' % corpus.__name__)
	print("name = %s" % corpus.__name__)
	#print "link = "
	ppath=''
	if hasattr(corpus,'PATH_TXT'):
		ppath=corpus.PATH_TXT
		print("path_txt = %s" % corpus.PATH_TXT)
	if hasattr(corpus,'PATH_XML'):
		if not ppath: ppath=corpus.PATH_XML
		print("path_xml = %s" % corpus.PATH_XML)
	if hasattr(corpus,'PATH_METADATA'): print("path_metadata = %s" % corpus.PATH_METADATA)
	print("path_python = %s" % ppath.split('/')[0] + '/' + ppath.split('/')[0] + '.py')
	print("class_corpus = %s" % corpus.__name__)
	print("class_text = %s" % 'Text'+corpus.__name__)


def do_configs(rootdir):
	import imp,os
	done=set()
	for fldr in sorted(os.listdir(rootdir)):
		path=os.path.join(rootdir,fldr)
		if not os.path.isdir(path): continue
		for fn in sorted(os.listdir(path)):
			if fn.endswith('.py') and not fn.startswith('_'):

				mod = imp.load_source(fn.replace('.py',''),os.path.join(path,fn))

				for obj in dir(mod):
					if obj[0]==obj[0].upper() and not obj in ['Text','Corpus'] and not obj.startswith('Text'):
						if obj in done: continue
						done|={obj}
						x=getattr(mod,obj)
						if not hasattr(x,'__name__'): continue
						print_config(x)


def gleanPunc2(aToken):
	aPunct0 = ''
	aPunct1 = ''
	while(len(aToken) > 0 and not aToken[0].isalnum()):
		aPunct0 = aPunct0+aToken[:1]
		aToken = aToken[1:]
	while(len(aToken) > 0 and not aToken[-1].isalnum()):
		aPunct1 = aToken[-1]+aPunct1
		aToken = aToken[:-1]

	return (aPunct0, aToken, aPunct1)

def modernize_spelling_in_txt(txt,spelling_d):
	lines=[]
	for ln in txt.split('\n'):
		ln2=[]
		for tok in ln.split(' '):
			p1,tok,p2=gleanPunc2(tok)
			tok=spelling_d.get(tok,tok)
			ln2+=[p1+tok+p2]
		ln2=' '.join(ln2)
		lines+=[ln2]
	return '\n'.join(lines)


def tokenize_fast(line):
	return re.findall("[A-Z]{2,}(?![a-z])|[A-Z][a-z]+(?=[A-Z])|[\'\w\-]+",line.lower())





### multiprocessing
def crunch(objects,function_or_methodname,ismethod=None,nprocs=8,args=[],kwargs={}):
	import time,random,six
	#ismethod=type(function_or_methodname) in [str,six.text_type] if ismethod is None else ismethod
	ismethod=type(function_or_methodname) in [str] if ismethod is None else ismethod

	def do_preparse(text,args=[],kwargs={}):
		threadid=os.getpid()
		time.sleep(random.uniform(0,5))
		print("[{2}] Starting working on {0} at {1}".format(text if False else 'ObjectX', now(), threadid))
		#print ismethod,function_or_methodname,args,kwargs
		if ismethod:
			x=getattr(text,function_or_methodname)(*args,**kwargs)
		else:
			x=function_or_methodname(text, *args, **kwargs)

		print("[{2}] Finished working on {0} at {1}".format(text if False else 'ObjectX', now(), threadid))
		return x

	import six.moves._thread,multiprocessing,os
	from multiprocessing import Process, Pipe
	#from itertools import zip
	izip=zip

	def spawn(f):
		def fun(q_in,q_out):
			numdone=0
			while True:
				numdone+=1
				i,x = q_in.get()
				if i == None:
					break
				q_out.put((i,f(x,args=args,kwargs=kwargs)))
		return fun

	def parmap(f, X, nprocs = multiprocessing.cpu_count()):
		q_in   = multiprocessing.Queue(1)
		q_out  = multiprocessing.Queue()

		proc = [multiprocessing.Process(target=spawn(f),args=(q_in,q_out)) for _ in range(nprocs)]
		for p in proc:
			p.daemon = True
			p.start()

		sent = [q_in.put((i,x)) for i,x in enumerate(X)]
		[q_in.put((None,None)) for _ in range(nprocs)]
		res = [q_out.get() for _ in range(len(sent))]

		[p.join() for p in proc]

		return [x for i,x in sorted(res)]

	parmap(do_preparse, objects, nprocs=nprocs)
	return True




def bigrams(l):
	return ngram(l,2)

def ngram(l,n=3):
	grams=[]
	gram=[]
	for x in l:
		gram.append(x)
		if len(gram)<n: continue
		g=tuple(gram)
		grams.append(g)
		gram.reverse()
		gram.pop()
		gram.reverse()
	return grams





### PASSAGES


def get_word_window(text,numwords=100,go_backwards=False):
	import re
	spaces = [match.start() for match in re.finditer(re.compile('\s'), text)]
	spaces = list(reversed(spaces)) if go_backwards else spaces
	spaces = spaces[:numwords]
	return text[:spaces[-1]] if not go_backwards else text[spaces[-1]:]

def index(text,phrase,ignorecase=True):
	compiled = re.compile(phrase, re.IGNORECASE) if ignorecase else re.compile(phrase)
	passage_indices = [(match.start(), match.end()) for match in re.finditer(compiled, text)]
	return passage_indices

def passages(text,phrases=[],window=200,indices=None,ignorecase=True,marker='***'):
	txt_lower = text.lower()
	window_radius=int(window/2)
	for phrase in phrases:
		if phrase.lower() in txt_lower:
			if not indices: indices = index(text,phrase,ignorecase=ignorecase)

			for ia,ib in indices:
				pre,post=text[:ia],text[ib:]
				match = text[ia:ib]
				window=get_word_window(pre,window_radius,True) + marker+match+marker+get_word_window(post,window_radius,False)
				dx={'index':ia, 'index_end':ib, 'passage':window,'phrase':phrase}
				yield dx

write = write2


def splitkeepsep(s, sep):
	return reduce(lambda acc, elem: acc[:-1] + [acc[-1] + elem] if elem == sep else acc + [elem], re.split("(%s)" % re.escape(sep), s), [])










## Spelling
V2S = None
def variant2standard():
	global V2S
	if not V2S:
		V2S = dict((d['variant'],d['standard']) for d in tools.tsv2ld(SPELLING_VARIANT_PATH,header=['variant','standard','']))
	return V2S

def standard2variant():
	v2s=variant2standard()
	d={}
	for v,s in list(v2s.items()):
		if not s in d: d[s]=[]
		d[s]+=[v]
	return d



def phrase2variants(phrase):
	s2v=standard2variant()
	words = phrase.split()
	word_opts = [[s]+s2v[s] for s in words]
	word_combos = list(tools.product(*word_opts))
	phrase_combos = [' '.join(x) for x in word_combos]
	return phrase_combos
###




ENGLISH = None
def load_english():
	global ENGLISH
	print('>> loading english dictionary...')
	ENGLISH = set(codecs.open('/Dropbox/LITLAB/TOOLS/english.txt','r','utf-8').read().split('\n'))
	#ENGLISH = (eng - load_stopwords())
	return ENGLISH







def yank(text,tag,none=None):
	if type(tag)==type(''):
		tag=tagname2tagtup(tag)

	try:
		return text.split(tag[0])[1].split(tag[1])[0]
	except IndexError:
		return none


def tagname2tagtup(tagname):
	return ('<'+tagname+'>','</'+tagname+'>')




def product(*args):
	if not args:
		return iter(((),)) # yield tuple()
	return (items + (item,)
		for items in product(*args[:-1]) for item in args[-1])


def zfy(data):
	from scipy.stats import zscore
	return zscore(data)




load_stopwords = get_stopwords







def linreg(X, Y):
	from math import sqrt
	from numpy import nan, isnan
	from numpy import array, mean, std, random

	if len(X)<2 or len(Y)<2:
		return 0,0,0
	"""
	Summary
		Linear regression of y = ax + b
	Usage
		real, real, real = linreg(list, list)
	Returns coefficients to the regression line "y=ax+b" from x[] and y[], and R^2 Value
	"""


	if len(X) != len(Y):  raise ValueError('unequal length')
	N = len(X)
	Sx = Sy = Sxx = Syy = Sxy = 0.0
	for x, y in map(None, X, Y):
		Sx = Sx + x
		Sy = Sy + y
		Sxx = Sxx + x*x
		Syy = Syy + y*y
		Sxy = Sxy + x*y
	det = Sxx * N - Sx * Sx
	a, b = (Sxy * N - Sy * Sx)/det, (Sxx * Sy - Sx * Sxy)/det
	meanerror = residual = 0.0
	for x, y in map(None, X, Y):
		meanerror = meanerror + (y - Sy/N)**2
		residual = residual + (y - a * x - b)**2

	RR = 1 - residual/meanerror if meanerror else 1
	ss = residual / (N-2) if (N-2) else 0
	Var_a, Var_b = ss * N / det, ss * Sxx / det
	#print "y=ax+b"
	#print "N= %d" % N
	#print "a= %g \\pm t_{%d;\\alpha/2} %g" % (a, N-2, sqrt(Var_a))
	#print "b= %g \\pm t_{%d;\\alpha/2} %g" % (b, N-2, sqrt(Var_b))
	#print "R^2= %g" % RR
	#print "s^2= %g" % ss
	return a, b, RR


def download_wget(url, save_to):
	import wget
	save_to_dir,save_to_fn=os.path.split(save_to)
	if save_to_dir: os.chdir(save_to_dir)
	fn=wget.download(url)#,bar=wget.bar_thermometer)
	os.rename(fn,save_to_fn)
	print('\n>> saved:',save_to)

def download(url,save_to):
	# ValueError: unknown url type: '%22https%3A//www.dropbox.com/s/wz3igeqzx3uu5j1/markmark.zip?dl=1"'
	url='https://' + url.split('//',1)[-1].replace('"','')
	return download_wget(url,save_to)

def download_tqdm(url, save_to):
	import requests
	from tqdm import tqdm

	r = requests.get(url, stream=True)
	total_size = int(r.headers.get('content-length', 0))

	with open(save_to, 'wb') as f:
		for chunk in tqdm(r.iter_content(32*1024), total=total_size, unit='B',unit_scale=True):
			if chunk:
				f.write(chunk)

	return save_to

def unzip(zipfn, dest='.'):
	from zipfile import ZipFile
	from tqdm import tqdm

	# Open your .zip file
	with ZipFile(zipfn) as zip_file:
		namelist=zip_file.namelist()

		# Loop over each file
		for file in tqdm(iterable=namelist, total=len(namelist)):
			# Extract each file to another directory
			# If you want to extract to current working directory, don't specify path
			zip_file.extract(member=file, path=dest)


def get_num_lines(filename):
	from smart_open import open

	def blocks(files, size=65536):
		while True:
			b = files.read(size)
			if not b: break
			yield b

	with open(filename, 'r', errors='ignore') as f:
		numlines=sum(bl.count("\n") for bl in blocks(f))

	return numlines



#print('>>>>',config)


def cloud_list(tmpfn='.tmp_llp_cloud_list'):
	import subprocess
	try:
		#out=subprocess.check_output(config['PATH_CLOUD_LIST_CMD'],shell=True)
		os.system(config['PATH_CLOUD_LIST_CMD'] +' > '+tmpfn)
		with open(tmpfn) as f: txt = f.read()
		os.unlink(tmpfn)
		return txt
	except Exception:
		return ''

def cloud_share_all():
	sharecmd=config['CLOUD_SHARE_CMD']
	dest=config['CLOUD_DEST']
