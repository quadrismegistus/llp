import codecs,configparser,os
LIT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
CONFIG_PATH = os.path.join(LIT_ROOT,'config.txt')
config = configparser.ConfigParser()
config.read(CONFIG_PATH)
config = dict([(k.upper(),v) for k,v in config['Default'].items()])

ENGLISH=None

def get_stopwords(include_rank=None):
	STOPWORDS_PATH = config.get('PATH_TO_ENGLISH_STOPWORDS')
	if not STOPWORDS_PATH: raise Exception('!! PATH_TO_ENGLISH_STOPWORDS not set in config.txt')
	if not STOPWORDS_PATH.startswith(os.path.sep): STOPWORDS_PATH=os.path.join(LIT_ROOT,STOPWORDS_PATH)
	sw1=set(codecs.open(STOPWORDS_PATH,encoding='utf-8').read().strip().split('\n'))
	if include_rank and type(include_rank)==int:
		sw2={d['word'] for d in worddb() if int(d['rank'])<=include_rank}
		sw1|=sw2
	return sw1

def get_english_wordlist():
	ENG_PATH = config.get('PATH_TO_ENGLISH_WORDLIST')
	if not ENG_PATH: raise Exception('!! PATH_TO_ENGLISH_WORDLIST not set in config.txt')
	if not ENG_PATH.startswith(os.path.sep): ENG_PATH=os.path.join(LIT_ROOT,ENG_PATH)
	print 'loading english from %s' % ENG_PATH
	return set(codecs.open(ENG_PATH,encoding='utf-8').read().strip().split('\n'))

def get_spelling_modernizer():
	SPELLING_MODERNIZER_PATH = config.get('PATH_TO_ENGLISH_SPELLING_MODERNIZER')
	if not SPELLING_MODERNIZER_PATH: raise Exception('!! PATH_TO_ENGLISH_SPELLING_MODERNIZER not set in config.txt')
	if not SPELLING_MODERNIZER_PATH.startswith(os.path.sep): SPELLING_MODERNIZER_PATH=os.path.join(LIT_ROOT,SPELLING_MODERNIZER_PATH)

	print '>> getting spelling modernizer from %s...' % SPELLING_MODERNIZER_PATH
	d={}
	with codecs.open(SPELLING_MODERNIZER_PATH,encoding='utf-8') as f:
		for ln in f:
			ln=ln.strip()
			if not ln: continue
			try:
				old,new=ln.split('\t')
			except ValueError:
				continue
			d[old]=new
	return d


def measure_ocr_accuracy(txt_or_tokens):
	global ENGLISH
	if type(txt_or_tokens) in [str,unicode]:
		tokens=tokenize(txt_or_tokens)
	elif type(txt_or_tokens) in [tuple,list]:
		tokens=list(txt_or_tokens)
	else:
		raise Exception("Function `measure_ocr_accuracy(txt_or_tokens)` must take text string or list of tokens.")

	if not ENGLISH: ENGLISH = get_english_wordlist()

	numwords=float(len(tokens))
	numenglish=len([tok for tok in tokens if tok in ENGLISH])
	return numenglish/numwords


def tokenize(txt):
	from nltk import word_tokenize
	return word_tokenize(txt)



def find_nth_character(str1, substr, n):
	pos = -1
	for x in xrange(n):
		pos = str1.find(substr, pos+1)
		if pos == -1:
			return None
	return pos

## only singular nouns!
def to_singular(ld):
	import inflect
	p=inflect.engine()
	return [d for d in ld if p.singular_noun(d['word']) in {d['word'],False}]

def worddb(abs_key = 'Complex Substance (Locke) <> Mixed Modes (Locke)_max',conc_key='Complex Substance (Locke) <> Mixed Modes (Locke)_min',cutoff_abs=0.1,cutoff_conc=-0.1,allow_names=False,only_content_words=True):
	worddb = read_ld('/Users/ryan/DH/18C/data/data.worddb.txt')
	for d in worddb:
		d['Abstract/Concrete'] = ''

		abs_score = float(d[abs_key])
		conc_score = float(d[conc_key])
		if only_content_words and d['is_content_word']!='True': continue
		if not allow_names and d['is_name']=='True': continue

		if abs_score >= cutoff_abs:
			d['Abstract/Concrete'] = 'Abstract'
		elif conc_score <= cutoff_conc:
			d['Abstract/Concrete'] = 'Concrete'
		else:
			d['Abstract/Concrete'] = 'Neither'

	return worddb


###





def read_ld(fn,keymap={},toprint=True):
	if fn.endswith('.xls') or fn.endswith('.xlsx'):
		return xls2ld(fn,keymap=keymap)
	elif fn.endswith('.csv'):
		#return tsv2ld(fn,tsep=',',keymap=keymap)
		tsep=','
		import csv
		with codecs.open(fn,encoding='utf-8') as f:
			return list(csv.DictReader(f))
	else:
		tsep='\t'

	return list(readgen(fn,tsep=tsep,as_dict=True,toprint=toprint))
		#return tsv2ld(fn,keymap=keymap)



def writegen(fnfn,generator,header=None,args=[],kwargs={}):
	of = codecs.open(fnfn,'w',encoding='utf-8')
	for i,dx in enumerate(generator(*args,**kwargs)):
		if not header: header=sorted(dx.keys())
		if not i: of.write('\t'.join(header) + '\n')
		of.write('\t'.join([unicode(dx.get(h,'')) for h in header]) + '\n')


def writegengen(fnfn,generator,header=None,save=True):
	if save: of = codecs.open(fnfn,'w',encoding='utf-8')
	for dx in generator():
		if not header:
			header=sorted(dx.keys())
			if save: of.write('\t'.join(header) + '\n')
		if save: of.write('\t'.join([unicode(dx.get(h,'')) for h in header]) + '\n')
		yield dx

def readgen(fnfn,header=None,tsep='\t',keymap={},keymap_all=unicode,encoding='utf-8',as_list=False,as_tuples=False,as_dict=True,toprint=True):
	if tsep=='\t' and toprint:
		print '>> streaming as tsv:',fnfn
	elif tsep==',' and toprint:
		print '>> streaming as csv:',fnfn
	import time
	now=time.time()
	header=None
	if fnfn.endswith('.gz'):
		import gzip
		of=gzip.open(fnfn)
	#of = codecs.open(fnfn,encoding=encoding)
	else:
		of=open(fnfn)

	for line in of:
		line=line.decode(encoding=encoding)[:-1]
		if not header:
			header=line.split(tsep)
			continue

		r=data=line.split(tsep)
		if as_list:
			yield r
			continue

		if as_tuples or as_dict:
			r=tuples=zip(header,data)
		if as_dict:
			r=d=dict(tuples)
		yield r
	of.close()
	nownow=time.time()
	if toprint: print '   done ['+str(round(nownow-now,1))+' seconds]'

def header(fnfn,tsep='\t',encoding='utf-8'):
	header=[]

	if fnfn.endswith('.gz'):
		import gzip
		of=gzip.open(fnfn)
	#of = codecs.open(fnfn,encoding=encoding)
	else:
		of=open(fnfn)

	for line in of:
		line = line[:-1]  # remove line end character
		line=line.decode(encoding=encoding)
		header=line.split(tsep)
		break
	of.close()
	return header

def read(fnfn,to_unicode=True):
	if fnfn.endswith('.gz'):
		import gzip
		try:
			with gzip.open(fnfn,'rb') as f:
				x=f.read()
				if to_unicode: x=x.decode('utf-8')
				return x
		except IOError as e:
			print "!! error:",e,
			print "!! opening:",fnfn
			print
			return ''

	elif fnfn.endswith('.txt'):
		if to_unicode:
			try:
				with codecs.open(fnfn,encoding='utf-8') as f:
					return f.read()
			except UnicodeDecodeError:
				return read(fnfn,to_unicode=False)
		else:
			with open(fnfn) as f:
				return f.read()

	return ''

def filesize(fn):
	return sizeof_fmt(os.path.getsize(fn))

def sizeof_fmt(num, suffix='B'):
	for unit in ['','Ki','Mi','Gi','Ti','Pi','Ei','Zi']:
		if abs(num) < 1024.0:
			return "%3.1f%s%s" % (num, unit, suffix)
		num /= 1024.0
	return "%.1f%s%s" % (num, 'Yi', suffix)



def xls2ld(fn,header=[],sheetname=True,keymap={},keymap_all=unicode):
	import time
	now=time.time()
	print '>> reading as xls:',fn
	import xlrd
	if '*' in keymap: keymap_all=keymap['*']
	headerset=True if len(header) else False
	f=xlrd.open_workbook(fn)
	ld=[]
	def _boot_xls_sheet(sheet,header=[]):
		ld2=[]
		for y in range(sheet.nrows):
			if not header:
				for xi in range(sheet.ncols):
					cell=sheet.cell_value(rowx=y,colx=xi)
					header+=[cell]
				continue
			d={}
			for key in header:
				try:
					value=sheet.cell_value(rowx=y, colx=header.index(key))
					#print '??',value,type(value),key
					if keymap_all:
						func=keymap_all
						if func in [str,unicode] and type(value) in [float]:
							if value == int(value): value=int(value)
						d[key]=keymap_all(value)
					elif keymap and key in keymap:
						func=keymap[key]
						if func in [str,unicode] and type(value) in [float]:
							if value == int(value): value=int(value)
						d[key]=keymap[key](value)
					else:
						d[key]=value
					#print key,value,y,header.index(key),row[header.index(key)]
				except Exception as e:
					print '!! ERROR:',e
					print '!! on key =',key,'& value =',value, type(value)
					#print "!! "+key+" not found in "+str(sheet)
					#d[key]=''
					pass
			if len(d):
				if sheetname: d['sheetname']=sheet.name
				ld2.append(d)
		return ld2


	if f.nsheets > 1:
		sheetnames=sorted(f.sheet_names())
		for sheetname in sheetnames:
			sheet=f.sheet_by_name(sheetname)
			for d in _boot_xls_sheet(sheet,header=header if headerset else []):
				ld.append(d)
	else:
		sheet = f.sheet_by_index(0)
		ld.extend(_boot_xls_sheet(sheet,header=header if headerset else []))

	nownow=time.time()
	print '>> done ['+str(round(nownow-now,1))+' seconds]'

	return ld


def xls2dld(fn,header=[]):
	return ld2dld(xls2ld(fn,header=header,sheetname=True), 'sheetname')

def levenshtein(s1, s2):
	l1 = len(s1)
	l2 = len(s2)

	matrix = [range(l1 + 1)] * (l2 + 1)
	for zz in range(l2 + 1):
		matrix[zz] = range(zz,zz + l1 + 1)
	for zz in range(0,l2):
		for sz in range(0,l1):
			if s1[sz] == s2[zz]:
				matrix[zz+1][sz+1] = min(matrix[zz+1][sz] + 1, matrix[zz][sz+1] + 1, matrix[zz][sz])
			else:
				matrix[zz+1][sz+1] = min(matrix[zz+1][sz] + 1, matrix[zz][sz+1] + 1, matrix[zz][sz] + 1)
	return matrix[l2][l1]

def xlsx2ld(fn,header=[],numsheets=1):
	from openpyxl.reader.excel import load_workbook
	header_set=bool(len(header))
	wb=load_workbook(filename=fn)
	ld=[]
	for sheet in wb.worksheets[:numsheets]:
		if not header_set: header=[]
		#header=[]
		for rownum,row in enumerate(sheet.rows):
			values=[]
			for cell in row:
				value=cell.value
				if value is None:
					value=''

				try:
					value=float(value)/0
				except:
					value=value
					if not isinstance(value, unicode):
						value=unicode(value)
				values.append(value)
			if not rownum and not len(header):
				header=values
			else:
				d=dict((header[i],values[i]) for i in range(len(values)))
				ld+=[d]
	return ld

def dl2ld(dl,kcol='group'):
	ld=[]
	for k in dl:
		for d in dl[k]:
			d[kcol]=k
			ld+=[d]
	return ld

def ld2dl(ld):
	keys = ld[0].keys()
	dl={}
	for k in keys:
		dl[k] = [d[k] for d in ld]
	return dl

def fn2ld(fn,header=[],sep='\t',nsep='\n'):
	import codecs
	f=codecs.open(fn,encoding='utf-8')
	for line in f:
		line=line.strip()
		if not header:
			header=line.split(sep)
			continue
		dx={}
		for i,val in enumerate(line.split(sep)):
			key=header[i] if len(header)>i else 'key_'+str(i)
			dx[key]=val
		yield dx

def goog2tsv(googsrc):
	import bs4
	dom=bs4.BeautifulSoup(googsrc,'html.parser')
	header=[th.text for th in dom('thead')[0]('th')]
	header=header if True in [bool(hx) for hx in header] else None
	old=[]
	for row in dom('tbody')[0]('tr'):
		rowdat=[cell.text for cell in row('td')]
		if not header:
			header=rowdat
			#print ">> HEADER:",header
			continue
		odx=dict(zip(header,rowdat))
		old+=[odx]
	return old


def tsv2ld(fn,tsep='\t',nsep='\n',u=True,header=[],keymap={},zero='',removeEmpties=False):
	import time
	now=time.time()
	if tsep=='\t':
		print '>> reading as tsv:',fn
	elif tsep==',':
		print '>> reading as csv:',fn

	import os
	if fn.startswith('http'):
		print '>> reading webpage...'
		import urllib
		f=urllib.urlopen(fn)
		t=f.read()
		if fn.endswith('/pubhtml'):
			return goog2tsv(t)
		f.close()
	elif not os.path.exists(fn):
		t=fn
	elif u:
		import codecs
		f=codecs.open(fn,encoding='utf-8')
		t=f.read()
		f.close()
	else:
		f=open(fn,'r')
		t=f.read()
		f.close()
	t=t.replace('\r\n','\n')
	t=t.replace('\r','\n')

	#header=[]
	listdict=[]


	for line in t.split(nsep):
		if not line.strip(): continue
		line=line.replace('\n','')
		ln=line.split(tsep)
		#print ln
		if not header:
			header=ln
			for i,v in enumerate(header):
				if v.startswith('"') and v.endswith('"'):
					header[i]=v[1:-1]
			continue
		edict={}
		for i in range(len(ln)):
			try:
				k=header[i]
			except IndexError:
				#print "!! unknown column for i={0} and val={1}".format(i,ln[i])
				continue
			v=ln[i].strip()

			if '*' in keymap:
				v=keymap['*'](v)
			elif k in keymap:
				#print v, type(v)
				v=keymap[k](v)
				#print v, type(v)
			else:
				if v.startswith('"') and v.endswith('"'):
					v=v[1:-1]
				try:
					v=float(v)
				except ValueError:
					v=v

			if type(v) in [str,unicode] and not v:
				if zero=='' and removeEmpties:
					continue
				else:
					v=zero
			edict[k]=v
		if edict:
			listdict.append(edict)

	nownow=time.time()
	print '>> done ['+str(round(nownow-now,1))+' seconds]'

	return listdict



def ld2html(ld):
	keys=ld2keys(ld)
	headerrow=['<th>%s</th>'%k for k in keys]
	rows=[]
	rows+=['\n\t\t'.join(headerrow)]
	for d in ld:
		row=['<td>%s</td>'%d.get(k,'') for k in keys]
		rows+=['\n\t\t'.join(row)]
	ostr=u"<table>\n\t<tr>\n\t\t" + u'\n\t</tr>\n\t<tr>\n\t\t'.join(rows) + u"\n\t</tr>\n</table>"
	return ostr

def ld2keys(ld):
	keys=[]
	for d in ld:
		for k in d:
			keys+=[k]
	keys=list(sorted(list(set(keys))))
	return keys

def ld2ll(ld,zero='',tostr=False,uni=True):
	keys=[]
	for d in ld:
		for k in d:
			keys+=[k]
	keys=sorted(list(set(keys)))
	o=[keys]
	for d in ld:
		l=[]
		for k in keys:
			v=d.get(k,zero)
			if tostr:
				v=unicode(v) if uni else str(v)
			l+=[v]
		o+=[l]
	return o


def write_ld(fn,ld,zero='',timestamp=None):
	return write(fn,ld2ll(ld,zero=zero),timestamp=timestamp)

def dd2ld(dd,rownamecol='rownamecol'):
	if not rownamecol:
		return [ (dict(v.items())) for k,v in dd.items() ]
	else:
		return [ (dict(v.items() + [(rownamecol,k)])) for k,v in dd.items() ]

def dld2ld(dld,key='rownamecol'):
	ld=[]
	for k in dld:
		for d in dld[k]:
			d[key]=k
			ld+=[d]
	return ld

def ld_resample(ld,key='rownamecol',n=None):
	import random
	dld=ld2dld(ld,key)
	minlen_actually=min([len(dld[k]) for k in dld])
	minlen=minlen_actually if not n or n>minlen_actually else n
	ld2=[]
	print '>> resampling to minimum length of:',minlen
	for k in sorted(dld):
		print '>>',k,len(dld[k]),'-->',minlen
		ld2+=random.sample(dld[k],minlen)
	return ld2

def ld2dld(ld,key='rownamecol'):
	dld={}
	for d in ld:
		if not d[key] in dld: dld[d[key]]=[]
		dld[d[key]]+=[d]
	return dld

def ld2dd(ld,rownamecol='rownamecol'):
	dd={}
	for d in ld:
		dd[d[rownamecol]]=d
		#del dd[d[rownamecol]][rownamecol]
	return dd

def datatype(data,depth=0,v=False):
	def echo(dt):
		if not v: return
		for n in range(depth): print "\t",
		print '['+dt[0]+']'+dt[1:],
		try:
			print "[{0} records]".format(len(data),dt)
		except:
			print

	if type(data) in [str,unicode]:
		echo('string')
		return 's'
	elif type(data) in [float,int]:
		echo('number')
		return 'n'
	elif type(data) in [list]:
		echo('list')
		if not len(data):
			return 'l'
		else:
			return 'l'+datatype(data[0],depth=depth+1,v=v)
	elif type(data) in [dict]:
		echo('dictionary')
		if not len(data):
			return 'd'
		else:
			return 'd'+datatype(data.values()[0],depth=depth+1,v=v)
	else:
		#print "WHAT TYPE OF DATA IS THIS:"
		#print data
		#print type(data)
		#print
		return '?'


def limcols(ld,limcol=255):
	keyd={}
	keys=set()
	for d in ld:
		dkeys=set(d.keys())
		for key in dkeys-keys:
			keyd[key]=0
		keys|=dkeys
		for k in d:
			if d[k]:
				keyd[k]+=1

	cols=set(sorted(keyd.keys(), key=lambda _k: (-keyd[_k],_k))[:limcol])

	for d in ld:
		dkeys=set(d.keys())
		for key in dkeys-cols:
			del d[key]

	return ld

def ld2str(ld,**data):
	if data['limcol']:
		print ">> limiting columns"
		limcol=data['limcol']
		ld=limcols(ld,limcol)
	if 'limcol' in data:
		del data['limcol']
	return ll2str(ld2ll(ld),**data)

def d2ll(d):
	try:
		return [[k,v] for k,v in sorted(d.items(),key=lambda lt: -lt[1])]
	except:
		return [[k,v] for k,v in d.items()]

def d2str(d,uni=True):
	return ll2str(d2ll(d),uni=uni)

def strmake(x,uni=True):
	if uni and type(x) in [unicode]:
		return x
	elif uni and type(x) in [str]:
		return x.decode('utf-8',errors='replace')
	elif uni:
		return unicode(x)
	elif not uni and type(x) in [str]:
		return x
	elif not uni and type(x) in [unicode]:
		return x.encode('utf-8',errors='replace')

	print [x],type(x)
	return str(x)


def ll2str(ll,uni=True,join_line=u'\n',join_cell=u'\t'):
	if not uni:
		join_line=str(join_line)
		join_cell=str(join_cell)
		quotechar='"' if join_cell==',' else ''
	else:
		quotechar=u'"' if join_cell==',' else u''

	for line in ll:
		lreturn=join_cell.join([quotechar+strmake(cell,uni=uni)+quotechar for cell in line])+join_line
		yield lreturn

def l2str(l,uni=True,join_line=u'\n',join_cell=u'\t',quotechar=''):
	for line in l: yield strmake(line)+join_line

def write_ld2(fn,gen1,gen2,uni=True,badkeys=[]):
	def find_keys(gen):
		keys=set()
		for d in gen:
			keys=keys|set(d.keys())
		keys=keys-set(badkeys)
		return keys

	keys=list(sorted(list(find_keys(gen1))))
	numk=len(keys)

	import codecs
	of=codecs.open(fn,'w',encoding='utf-8')
	of.write('\t'.join([strmake(x) for x in keys]) + '\n')

	for d in gen2:
		data=[d.get(key,'') for key in keys]
		of.write('\t'.join([strmake(x) for x in data]) + '\n')
	of.close()
	print ">> saved:",fn


def write2(fn,data,uni=True,join_cell=u'\t',join_line=u'\n',limcol=None,toprint=True):
	## pass off to other write functions if necessary
	if fn.endswith('.xls'): return write_xls(fn,data)
	if fn.endswith('.csv'): join_cell=','

	## get datatyoe
	dt=datatype(data)

	## get str output for datatype
	if dt.startswith('ld'):
		o=ld2str(data,join_cell=join_cell,limcol=limcol)
	elif dt.startswith('dl'):
		o=dl2str(data,uni=uni)
	elif dt.startswith('ll'):
		o=ll2str(data,uni=uni)
	elif dt.startswith('dd'):
		o=dd2str(data,uni=uni)
	elif dt.startswith('l'):
		o=l2str(data,uni=uni)
	elif dt.startswith('d'):
		o=d2str(data,uni=uni)
	else:
		o=data

	## write
	import codecs
	of = codecs.open(fn,'w',encoding='utf-8') if True else open(fn,'w')
	for line in o: of.write(line)
	of.close()
	if toprint: print '>> saved:',fn


def now(now=None):
	import datetime as dt
	if not now:
		now=dt.datetime.now()
	elif type(now) in [int,float,str]:
		now=dt.datetime.fromtimestamp(now)

	return '{0}-{1}-{2} {3}:{4}:{5}'.format(now.year,str(now.month).zfill(2),str(now.day).zfill(2),str(now.hour).zfill(2),str(now.minute).zfill(2),str(now.second).zfill(2))



def toks2str(tlist,uni=False):
	toks=[]
	putleft=False
	#print tlist
	for tk in tlist:
		tk=tk.strip()
		if not tk: continue
		tk = tk.split()[-1]
		if not tk: continue
		if (not len(toks)):
			toks+=[tk]
		elif putleft:
			toks[-1]+=tk
			putleft=False
		elif tk=='`':
			toks+=[tk]
			putleft=True
		elif tk=='-LRB-':
			toks+=['(']
			putleft=True
		elif tk=='-RRB-':
			toks[-1]+=')'
		elif len(tk)>1 and tk[0]=="'":
			toks[-1]+=tk
		elif tk[0].isalnum():
			toks+=[tk]
		elif tk.startswith('<') and '>' in tk:
			toks+=[tk]
		else:
			toks[-1]+=tk
	if uni: return u' '.join(toks)
	return ' '.join(toks)






####
def print_config(corpus):
	print
	print
	print '[%s]' % corpus.__name__
	print "name = %s" % corpus.__name__
	#print "link = "
	ppath=''
	if hasattr(corpus,'PATH_TXT'):
		ppath=corpus.PATH_TXT
		print "path_txt = %s" % corpus.PATH_TXT
	if hasattr(corpus,'PATH_XML'):
		if not ppath: ppath=corpus.PATH_XML
		print "path_xml = %s" % corpus.PATH_XML
	if hasattr(corpus,'PATH_METADATA'): print "path_metadata = %s" % corpus.PATH_METADATA
	print "path_python = %s" % ppath.split('/')[0] + '/' + ppath.split('/')[0] + '.py'
	print "class_corpus = %s" % corpus.__name__
	print "class_text = %s" % 'Text'+corpus.__name__


def do_configs(rootdir):
	import imp,os
	done=set()
	for fldr in sorted(os.listdir(rootdir)):
		path=os.path.join(rootdir,fldr)
		if not os.path.isdir(path): continue
		for fn in sorted(os.listdir(path)):
			if fn.endswith('.py') and not fn.startswith('_'):

				mod = imp.load_source(fn.replace('.py',''),os.path.join(path,fn))

				for obj in dir(mod):
					if obj[0]==obj[0].upper() and not obj in ['Text','Corpus'] and not obj.startswith('Text'):
						if obj in done: continue
						done|={obj}
						x=getattr(mod,obj)
						if not hasattr(x,'__name__'): continue
						print_config(x)


def gleanPunc2(aToken):
	aPunct0 = ''
	aPunct1 = ''
	while(len(aToken) > 0 and not aToken[0].isalnum()):
		aPunct0 = aPunct0+aToken[:1]
		aToken = aToken[1:]
	while(len(aToken) > 0 and not aToken[-1].isalnum()):
		aPunct1 = aToken[-1]+aPunct1
		aToken = aToken[:-1]

	return (aPunct0, aToken, aPunct1)

def modernize_spelling_in_txt(txt,spelling_d):
	lines=[]
	for ln in txt.split('\n'):
		ln2=[]
		for tok in ln.split(' '):
			p1,tok,p2=gleanPunc2(tok)
			tok=spelling_d.get(tok,tok)
			ln2+=[p1+tok+p2]
		ln2=' '.join(ln2)
		lines+=[ln2]
	return '\n'.join(lines)






### multiprocessing
def crunch(objects,function_or_methodname,ismethod=None,nprocs=8,args=[],kwargs={}):
	import time,random
	ismethod=type(function_or_methodname) in [str,unicode] if ismethod is None else ismethod

	def do_preparse(text,args=[],kwargs={}):
		threadid=os.getpid()
		time.sleep(random.uniform(0,5))
		print "[{2}] Starting working on {0} at {1}".format(text if False else 'ObjectX', now(), threadid)
		#print ismethod,function_or_methodname,args,kwargs
		if ismethod:
			x=getattr(text,function_or_methodname)(*args,**kwargs)
		else:
			x=function_or_methodname(text, *args, **kwargs)

		print "[{2}] Finished working on {0} at {1}".format(text if False else 'ObjectX', now(), threadid)
		return x

	import thread,multiprocessing,os
	from multiprocessing import Process, Pipe
	from itertools import izip

	def spawn(f):
		def fun(q_in,q_out):
			numdone=0
			while True:
				numdone+=1
				i,x = q_in.get()
				if i == None:
					break
				q_out.put((i,f(x,args=args,kwargs=kwargs)))
		return fun

	def parmap(f, X, nprocs = multiprocessing.cpu_count()):
		q_in   = multiprocessing.Queue(1)
		q_out  = multiprocessing.Queue()

		proc = [multiprocessing.Process(target=spawn(f),args=(q_in,q_out)) for _ in range(nprocs)]
		for p in proc:
			p.daemon = True
			p.start()

		sent = [q_in.put((i,x)) for i,x in enumerate(X)]
		[q_in.put((None,None)) for _ in range(nprocs)]
		res = [q_out.get() for _ in range(len(sent))]

		[p.join() for p in proc]

		return [x for i,x in sorted(res)]

	parmap(do_preparse, objects, nprocs=nprocs)
	return True




def bigrams(l):
	return ngram(l,2)

def ngram(l,n=3):
	grams=[]
	gram=[]
	for x in l:
		gram.append(x)
		if len(gram)<n: continue
		g=tuple(gram)
		grams.append(g)
		gram.reverse()
		gram.pop()
		gram.reverse()
	return grams
